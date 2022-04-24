import openpyxl, sys

#file = sys.argv[1]
#file = input('SpreedSheet to invert Cells')
file = 'Items.xlsx'
wb = openpyxl.load_workbook(file)
sheet = wb.active
newSheet = wb.create_sheet(index = 0, title = 'inverted')
print(sheet.rows)
for x in sheet.rows:
    print(x)
    for y in x:
        print(y)
        newSheet.cell(row = y.column, column = y.row).value = y.value
#wb.save('inverted_{}'.format(file))

