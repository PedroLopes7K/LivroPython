import openpyxl
from openpyxl.styles import Font

def multiplicationTable(number):
    wb =openpyxl.Workbook()
    sheet=wb.active
    #CONGELA A CELULA A1
    #A linha 1 e a coluna A devem ser usadas para rótulos e devem estar em negrito.
    sheet.freeze_panes = 'A2'
    sheet.freeze_panes = 'B1'
    fontObj1 = Font(name='Times New Roman', bold=True)
    #Os dois primeiros for inserem os numeros de 1 ao numero passado na funçã, na vertical e horizontal
    for i in range(1,number+1):
        sheet.cell(row=i+1,column=1).value=i
        sheet.cell(row=i+1,column=1).font=fontObj1
    for i in range(1,number+1):
        sheet.cell(row=1,column=i+1).value=i
        sheet.cell(row=1,column=i+1).font=fontObj1
    # o terceiro for executa as multiplicações
    for row in range(1,number+1):
        for cell in range(1,number+1):
            sheet.cell(row=row+1,column=cell+1).value=row*cell
    wb.save('multiplicationTable'+str(number)+'.xlsx')
        
myNumber=int(input('Enter the number you want to create a multiplication table for :'))     
multiplicationTable(myNumber)

