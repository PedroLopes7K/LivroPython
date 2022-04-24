# FUNDIR E DESFUNDIR CELULAS
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3') # fundi todas essas celulas
# Para setar o valor das celulas fundidad, basta adicionar o valor a celula do topo a esquerda
sheet['A1'] = ' TODAS ESSAS CELULAS ESTÃO FUNDIDAS'
sheet.merge_cells('C5:E5')
sheet['C5'] = '3 celulas fundidas'
wb.save('FundirCelulas.xlsx')
print('SUCESS')

''' PARA DESFUNDIR CELULAS USA-SE unmerge

 import openpyxl
>>> wb = openpyxl.load_workbook('merged.xlsx')
>>> sheet = wb.active
>>> sheet.unmerge_cells('A1:D3') # Split these cells up.
>>> sheet.unmerge_cells('C5:D5')
>>> wb.save('merged.xlsx')

'''
 
