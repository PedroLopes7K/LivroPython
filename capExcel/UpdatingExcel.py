#Corrigir custos na planilha de vendas de produtos

import openpyxl

wb = openpyxl.load_workbook('Produtos.xlsx')
planilha = wb['Planilha1']

#os produtos e seus preços atualizados
PRECO_ATUALIZADO = {'Batatas': 2.07,
                    'Alho': 1.19,
                    'Limao': 1.27}
# A FAZER: PERCORRER AS LINHAS E ATUALIZAR OS PREÇOS

for numLinha in range(2, planilha.max_row): # pula a primeira linha com titulo das colunas
    NomeProduto = planilha.cell(row=numLinha, column=1).value
    if NomeProduto in PRECO_ATUALIZADO:
        planilha.cell(row=numLinha, column=2).value = PRECO_ATUALIZADO[NomeProduto]
wb.save('Produtos.xlsx') # salvando na mesma planilha
# wb.save('ProdutosAtualizados.xlsx') salvando em uma nova planilha
print('Sucess')
